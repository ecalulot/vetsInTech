# Group 5: Arthur, Erika, Geoffrey, Justin 
import json, requests
from openpyxl import Workbook, load_workbook


# list of zipcodes
zip_list = []

top_10_us_cities = [
    "New York",
    "Los Angeles",
    "Chicago",
    "Houston",
    "Phoenix",
    "Philadelphia",
    "San Antonio",
    "San Diego",
    "Dallas",
    "San Jose",
    "Austin",
    "Jacksonville",
    "Fort Worth",
    "Columbus",
    "Charlotte",
    "Indianapolis",
    "San Francisco",
    "Denver",
    "Seattle",
    "Orlando"
]
   

# loop to ask for 5 zipcode inputs
for i in range(1,6):
    zip = input("Enter zipcode: ")
    zip_list.append(zip)

print(zip_list)

wb = Workbook()
ws = wb.active
ws.title = "The Weather and You"

# header for excel worksheet
ws.cell(row=1, column=1, value='Zipcode')
ws.cell(row=1, column=2, value='City')

# populate excel worksheet with input zipcodes and it's equivalent city names
for i in range(1,len(zip_list)+1):
    ws.cell(row=i+1, column=1, value=zip_list[i-1])
    api_url = 'http://ZiptasticAPI.com/{}'.format(zip_list[i-1])
    response = requests.get(api_url)
    data = response.json()
    try: #check if zipcode is valid or if it exists
        ws.cell(row=i+1, column=2, value=data['city'])
    except KeyError: #displays invalid zipcode information and keeps the spreadsheet blank for city
        print("Zipcode is invalid: ", zip_list[i-1])
        ws.cell(row=i+1, column=2, value="Not Valid")

wb.save('./spreadsheets/project.xlsx')


#load city names to get weather
load_wb = load_workbook('./spreadsheets/project.xlsx')
load_ws = load_wb.active

city_names = []

for row in load_ws:
    c = row[1].value
    if c == 'City':
        continue
    else: 
        if c == "Not Valid": # checks if city column is blank and skips it to append to city_names
            continue
        else: city_names.append(c.title())

print(city_names)

ws1 = wb.create_sheet("City Temperatures")
ws1.cell(row=1, column=1, value='City')
ws1.cell(row=1, column=2, value='Temperature in Celsius')

new_list = list(set(top_10_us_cities + city_names))


city_temp = []
for i in range(1,len(new_list)+1): # goes through city names and gets the current temperature in F and C of a specified city
    wx_api_url = "https://wttr.in/{}?format=j1".format(new_list[i-1])
    response = requests.get(wx_api_url)
    data = response.json()
    print("The current temperature in " + new_list[i-1] + " is: " + data["current_condition"][0]['temp_C'] + "C")    
    city_temp.append({"city":new_list[i-1],"temp":int(data["current_condition"][0]['temp_C'])})
sorted_list = sorted(city_temp, key=lambda x: x['temp'])  
print(sorted_list)
    
for i in range(1,len(sorted_list)):
    ws1.cell(row=i+1, column=1, value=sorted_list[i-1]["city"])
    ws1.cell(row=i+1, column=2, value=sorted_list[i-1]["temp"])

wb.save('./spreadsheets/project.xlsx')