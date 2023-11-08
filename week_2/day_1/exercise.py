# ITP Week 2 Day 1 Exercise

# SCENARIO: You're a store owner receiving the inventory report for this month.
# You will receive the product order spreadsheet soon and it is easier to calculate your order
# if your inventory was also on a spreadsheet! Recreate the following spreadsheet with Python: 

# don't forget your appropriate imports.
# assign the title of the initial active sheet to "CURRENT_MONTH_INVENTORY"

#   product_name    product_id  max_amount      reorder_threshold   quantity
# 	oreo            2323        1000            300                 743
# 	coke            6545        500             100                 101
# 	pepsi	        3456        200             50                  137
# 	lays_chip	    4567        1500            500                 364
# 	pringles	    2134        2000            600                 120
# 	sour_worms	    2362        100             10                  85
# 	choco_cookies   0923	    200             25                  24
# 	donuts	        2786        200             25                  12
# 	hot_dogs	    6723        100             10                  39
# 	ice_cream	    9237        200             50                  234
# 	gum	            2092        3500            1000                1232
# 	pretzels        8246	    100             5                   11
# 	kit_kat	        9276        1000            250                 249

# TIP: create a list of each column (ie. product_names = [oreo, ...]) and use those to loop through :)

# save your file
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "CURRENT_MONTH_INVENTORY"

product_name = ['product name', 'oreo', 'coke', 'pepsi', 'lays chip', 'pringles',
                'sour worms', 'choco cookies', 'donuts', 'hot dogs', 'ice cream', 
                'gum', 'pretzels', 'kit kat']
# print(f"Product Name: {len(product_name)}")

for num in range(1, len(product_name)+1):
    ws.cell(row=num, column=1, value=product_name[num-1])

product_id = ['product id', '2323', '6545', '3456', '4567', '2134', '2362',
              '0923', '2786', '6723', '9237', '2092', '8246', '9276'] 
# print(f"Product ID: {len(product_id)}")
for num in range(1, len(product_id)+1):
    ws.cell(row=num, column=2, value=product_id[num-1])

# create the columns into a list
# include the title in the list

max_amount = ['max amount', 1000, 500, 200, 1500, 2000, 100, 200,
              200, 100, 200, 3500, 100, 1000]
# print(f"Max amount: {len(max_amount)}")
for num in range(1, len(max_amount)+1):
    ws.cell(row=num, column=3, value=max_amount[num-1])

reorder_threshold = ['reorder threshold', 300, 100, 50, 500, 600, 10, 
                     25, 25, 10, 50, 1000, 5, 250]
# print(f"Reorder threshold: {len(reorder_threshold)}")
for num in range(1, len(reorder_threshold)+1):
    ws.cell(row=num, column=4, value=reorder_threshold[num-1])

quantity = ['quantity', 743, 101, 137, 364, 120, 
            85, 24, 12, 39, 234, 1232, 11, 249]
# print(f"Quantity: {len(quantity)}")
for num in range(1, len(quantity)+1):
    ws.cell(row=num, column=5, value=quantity[num-1])

# make sure to assign data types i.e. 'str' and 'int', product_id is a str though
print(wb.sheetnames)
# wb.save("./spreadsheets/inventory.xlsx")
wb.save("week_2/spreadsheets/inventory.xlsx") # by relative pathing