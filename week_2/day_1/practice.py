# ITP Week 2 Day 1 (In-Class) Practice

# A1. from the appropriate library, import only the Workbook
from openpyxl import Workbook

# A2. Before anything, we need a workbook to work with..
wb = Workbook()
print(type(wb))

# ws1 = wb.create_sheet("Rugrats")
# ws2 = wb.create_sheet("Hey Arnold", 0) 
# 0 moves the worksheet tab to move to the very front and is the frirst sheet tab

# A3. We need to interact with a single worksheet.
ws = wb.active # default activates the first worksheet
ws.title = "W2D1-practice" # gives the worksheet a title
print(wb.sheetnames)

# for sheet in wb:
#     print(sheet.title)

# A4. assign the value of "First Name" to A1
a1 = ws['A1'] = "First Name"
# not necessary to assign to a variable

# A5. assign the value of "Last Name" to B1
# b1 = ws['B1'] = "Last Name"
ws.cell(row=1, column=2, value="Last Name")
# ws.cell(row=4, column=1, value='Whatever I want') # columns are numbered
# this is the same as cell A4

# colC = ws['C'] # gets all of column C
# row10 = ws[10] # gets all of row 10


# STOP HERE - RETURN TO LECTURE

# B1. For all of column A, starting at row 2 until row 10, make the cell values: "Gabriel" (attempt a loop)

# column_a = ws['A']
# for cell in column_a:
#     cell.value = 'Gabriel'

for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=1):
    for cell in row:
        cell.value = "Gabriel"

# last_names = ['Rolley', 'Smith', 'Balenga', 'Issac', 'Cruise', 'Depp', 'Heard', 'Qiao', 'Biden']

# # B2. Loop through a range from row 2 to 10 and assign the cell value to last names according to index in column B
# # NOTE: PAY ATTENTION to the starting number of the range and how it differs from the starting index of the list

# for row in ws.iter_rows(min_row=2, max_row=10, min_col=2): # rows 2-10, column B
#     for name in last_names:
#         cell.value = name
# # B3. Save the file
# wb.save("./spreadsheets/day_1_practice.xlsx")
wb.save("/home/ra11y/Downloads/local.source/vetsInTech/week_2/spreadsheets/day_1_practice.xlsx")