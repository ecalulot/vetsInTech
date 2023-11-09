# ITP Week 2 Day 2 Exercise

# import the appropriate method from the correct module
from openpyxl import Workbook, load_workbook
# Import the workbook that you updated in today's practice from
# "./spreadsheets/inventory.xlsx"

# relative path can run this program from within VS Code (ctrl+F5)
# wb = load_workbook("week_2/spreadsheets/inventory.xlsx") 

# absolute path can allow this program to run from terminal: "$python3 exercise.py"
wb = load_workbook("/home/ra11y/Downloads/local.source/vetsInTech/week_2/spreadsheets/inventory.xlsx")
# access and store the appropriate worksheet to the variable 'ws'
sheet_name_list = wb.sheetnames
print(f"\nThe sheet(s) in the workbook is called: {sheet_name_list}\n")
print("Setting the active sheet name... \n")
ws = wb['CURRENT_MONTH_INVENTORY']
m_row = ws.max_row 
print(f"The maximum rows in the sheet are: {m_row}\n")

# initialize list arrays to iterate through the Excel values
quantity = []
threshold = []
max_amt = []

# TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness

for rows in range(2, m_row+1):
# start at 2 to account for headers, m_row+1 accounts for python inherent "end/stop_value - 1"
# append values to their respective lists
    quantity.append(ws.cell(row=rows, column=5).value)
    threshold.append(ws.cell(row=rows, column=4).value)
    max_amt.append(ws.cell(row=rows, column=3).value)

# Define a function called add_order_amount that takes in a single parameter called 'row'
def add_order_amount(rows):
   # must subtract 2 to get to index[0] in initialized Python lists (lines 22-24)
    if quantity[rows-2] <= threshold[rows-2]:
        order_amount = max_amt[rows-2] - quantity[rows-2]
        ws.cell(row=rows, column=6, value=order_amount) # by openpyxl column=6 is 'F'
    else:
        ws.cell(row=rows, column=6, value="Skip reorder")
    # without this 'else' statement program was calculating the value from the very last cells, \
    # it outputting '751' from kit kat cell values: 1000 - 249
    
    # IF the quantity is less or equal to than the threshold,
        # than calculate the order_amount (max_amount - quantity)
        # assign the value to that row, column 6

# Perform a for..in loop through the range(2, len(inventory.rows))
#   - call the function add_order_amount() passing in the number of the range

for rows in range(2, (ws.max_row)+1):
    add_order_amount(rows)

wb.save("week_2/spreadsheets/inventory.xlsx")
# got to save xlsx file to see alterations. 