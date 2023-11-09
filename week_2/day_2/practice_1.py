# ITP Week 2 Day 2 (In-Class) Practice 1
# 
# You will continue to work on the inventory spreadsheet that you created from yesterday's exercise
# import the appropriate method from the correct module
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Import the workbook that you created in yesterday's exercise from
# "./spreadsheets/inventory.xlsx"
wb = load_workbook("week_2/spreadsheets/inventory.xlsx")

# verify what sheet names are available
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

ws = wb['CURRENT_MONTH_INVENTORY']
# # access and store the appropriate worksheet to the variable 'ws'


# # Print out the cell values for each row
# cell_val_range = ['A1':'E14']
# print(cell_val_range)


# all_rows = ws.rows
for row in ws.rows:
    for each_cell in row:
        print(each_cell.value)

# ws["F1"] = 'order_amount'

ws.cell(row=1, column=6, value='order amount')
# Create a new column within that worksheet called order_amount
# wb.insert_cols = ['len(max_column(wb)) + 1']
# save the latest changes
wb.save("week_2/spreadsheets/inventory.xlsx")