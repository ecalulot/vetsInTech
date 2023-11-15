# ITP Week 3 Day 1 Practice

# import your required modules/methods
from openpyxl import Workbook, load_workbook
# given the following items, using the methods we covered, write to openpyxl
wb = Workbook()
ws = wb.active
# ws.title = 'Demo Worksheet'

# use an external counter with just a for loop (no function)

clefairy = {
    "id": 35,
    "name": "clefairy",
    "base_experience": 113,
    "height": 6,
    "order": 56,
    "weight": 75,
    }

# print(len(clefairy))
for num in range(1, len(clefairy.keys())+1):
    ws.cell(row=1, column=num, value=list(clefairy.values())[num-1])
# [num-1] is to start at index 0 when we convert the clefairy dictionary into a list so we can iterate through it.

# # using an external counter method
# counter = 1
# for key in clefairy:
#     ws.cell(row=6, column=counter, value=clefairy[key] # we're not using the .key() method we are indexing it
#     counter += 1

# create a function that takes in a pokemon
weedle = {
    "id": 13,
    "name": "weedle",
    "base_experience": 39,
    "height": 3,
    "order": 17,
    "weight": 32,
    }

# call the function with weedle!
def get_pokemon(weedle):
    for col in range(1, len(weedle.keys())+1):
        ws.cell(row=2, column=col, value=list(weedle.values())[col-1])
     
get_pokemon(weedle)
# wb.save('./spreadsheets/practice.xlsx')
# absolute path
wb.save('/home/ra11y/Downloads/local.source/vetsInTech/week_3/spreadsheets/w3_d1_practice.xlsx')