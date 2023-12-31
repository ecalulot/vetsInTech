# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)

# EASY MODE

# import the appropriate modules (you have 3)
from openpyxl import Workbook, load_workbook
import requests, json

multi_character_url = "https://rickandmortyapi.com/api/character"

single_character_url = "https://rickandmortyapi.com/api/character/6" # for a single character
# set up a workbook and worksheet titled "Rick and Morty Characters"
wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"

# # assign a variable 'data' with the returned GET request
url_data_single = requests.get(single_character_url)
print(url_data_single) # 200 if website is available
# print(type(data))
prep_data = url_data_single.text
json_data_single = json.loads(prep_data) # converts json to python dictionary and/or lists
print(f"'json_data' is of type: {type(json_data_single)}")
# print(json_data)
json_str= json.dumps(json_data_single, indent=4) # convert to json str obj to see structure of the data
# print(json_str)

# data/url/api to get information to fill the remainder of the xlsx 
url_data_multi = requests.get(multi_character_url)
prep_data_multi = url_data_multi.text
json_data_multi = json.loads(prep_data_multi)


# create the appropriate headers in openpyxl for all of the keys for a single character
for col, post in enumerate(json_data_single, 1): # this creates the headers for a single character, url (line 15) adjusted to get one character. 
    ws.cell(row=1, column=col, value=post)
    for row, multi_data in enumerate(json_data_multi.items(), 1):
        ws.cell(row=2, column=row, value=str(multi_data[1])) # can call the index of a tuple, but not quite correct if there is a dictionary within


# loop through all of the 'results' of the data to populate the rows and columns for each character 
# def character_data(url_data): 
#     for col in range(1, len(data.keys())+1):
#         ws.cell(row=2, column=col, value=list(data.values())[col-1])

# character_data(data) # NOTE: due to the headers, the rows need to be offset by one!
# MEDIUM MODE
# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"
sheet2 = wb.create_sheet(title="Rick and Morty Locations")
sheet3 = wb.create_sheet(title="Rick and Morty Episodes")
print("The workbook sheet names are...")
print(wb.sheetnames) 

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
location_url = "https://rickandmortyapi.com/api/location"
episode_url = "https://rickandmortyapi.com/api/episode"

# populate the new worksheets appropriately with all of the data!
data_location = requests.get(location_url)
data_episode = requests.get(episode_url)

# NOTE: don't forget your headers!

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)

# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)


wb.save("week_3/spreadsheets/w3_d3_exercise.xlsx")
# relative path

# wb.save("/home/ra11y/Downloads/local.source/vetsInTech/week_3/spreadsheets/w3_d3_exercise.xlsx")
# absolute path